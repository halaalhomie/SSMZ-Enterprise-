from fastapi import APIRouter, Depends, Query, UploadFile, File, WebSocket, WebSocketDisconnect, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from typing import Optional
from uuid import UUID
from fastapi.responses import StreamingResponse
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook

from app.models.models import User, Product, StockTransaction, TransactionType
from app.db.database import get_db
from app.core.deps import get_current_user, require_owner, get_current_store_id
from app.models.models import User
from app.schemas.schemas import (
    LoginRequest, TokenResponse, RefreshRequest, ChangePasswordRequest,
    UserCreate, UserUpdate, UserOut,
    CategoryCreate, CategoryUpdate, CategoryOut,
    SupplierCreate, SupplierUpdate, SupplierOut,
    ProductCreate, ProductUpdate, ProductOut,
    StockInCreate, StockOutCreate, TransactionOut,
    AuditCreate, AuditOut,
    NoteCreate, NoteUpdate, NoteOut,
    NotificationOut, ActivityLogOut,
)
from app.services.auth_service import AuthService
from app.services.product_service import ProductService
from app.services.stock_service import StockService
from app.services.audit_service import AuditService
from app.services.dashboard_service import DashboardService
from app.services.activity_service import ActivityService, NotificationService
from app.websockets.manager import manager
from app.core.security import decode_token

import logging
logger = logging.getLogger(__name__)

router = APIRouter()


# ═══════════════════════════════════════════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/auth/login", response_model=TokenResponse, tags=["Auth"])
async def login(data: LoginRequest, db: AsyncSession = Depends(get_db)):
    return await AuthService.login(db, data)


@router.post("/auth/register", response_model=TokenResponse, tags=["Auth"])
async def register(data: UserCreate, store_name: str = Query(...), db: AsyncSession = Depends(get_db)):
    return await AuthService.register_owner(db, data, store_name)


@router.post("/auth/refresh", response_model=TokenResponse, tags=["Auth"])
async def refresh(data: RefreshRequest, db: AsyncSession = Depends(get_db)):
    return await AuthService.refresh_token(db, data.refresh_token)


@router.post("/auth/change-password", tags=["Auth"])
async def change_password(
    data: ChangePasswordRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await AuthService.change_password(db, current_user, data.current_password, data.new_password)


@router.get("/auth/me", response_model=UserOut, tags=["Auth"])
async def me(current_user: User = Depends(get_current_user)):
    return current_user


# ═══════════════════════════════════════════════════════════════════════════════
# USERS  (owner only)
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/users", tags=["Users"])
async def list_users(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    current_user: User = Depends(require_owner),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select, func, or_
    from app.models.models import User as UserModel
    query = select(UserModel).where(UserModel.store_id == current_user.store_id)
    if search:
        query = query.where(or_(UserModel.name.ilike(f"%{search}%"), UserModel.email.ilike(f"%{search}%")))
    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar()
    users = (await db.execute(query.offset((page - 1) * page_size).limit(page_size))).scalars().all()
    return {"items": [UserOut.model_validate(u) for u in users], "total": total, "page": page, "page_size": page_size}


@router.post("/users", response_model=UserOut, tags=["Users"])
async def create_user(
    data: UserCreate,
    current_user: User = Depends(require_owner),
    db: AsyncSession = Depends(get_db),
):
    from app.models.models import User as UserModel
    from app.core.security import hash_password
    from sqlalchemy import select
    existing = (await db.execute(select(UserModel).where(UserModel.email == data.email))).scalar_one_or_none()
    if existing:
        from fastapi import HTTPException, status
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Email already exists")
    user = UserModel(store_id=current_user.store_id, name=data.name, email=data.email, password_hash=hash_password(data.password), role=data.role)
    db.add(user)
    await db.flush()
    return UserOut.model_validate(user)


@router.patch("/users/{user_id}", response_model=UserOut, tags=["Users"])
async def update_user(
    user_id: UUID,
    data: UserUpdate,
    current_user: User = Depends(require_owner),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select
    from app.models.models import User as UserModel
    from fastapi import HTTPException, status
    result = await db.execute(select(UserModel).where(UserModel.id == user_id, UserModel.store_id == current_user.store_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    for field, value in data.model_dump(exclude_none=True).items():
        setattr(user, field, value)
    return UserOut.model_validate(user)


# ═══════════════════════════════════════════════════════════════════════════════
# CATEGORIES
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/categories", tags=["Categories"])
async def list_categories(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select
    from app.models.models import Category
    result = await db.execute(select(Category).where(Category.store_id == current_user.store_id, Category.is_active == True))
    return [CategoryOut.model_validate(c) for c in result.scalars().all()]


@router.post("/categories", response_model=CategoryOut, tags=["Categories"])
async def create_category(
    data: CategoryCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.models.models import Category
    cat = Category(store_id=current_user.store_id, **data.model_dump())
    db.add(cat)
    await db.flush()
    return CategoryOut.model_validate(cat)


@router.patch("/categories/{cat_id}", response_model=CategoryOut, tags=["Categories"])
async def update_category(
    cat_id: UUID,
    data: CategoryUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select
    from app.models.models import Category
    from fastapi import HTTPException, status
    result = await db.execute(select(Category).where(Category.id == cat_id, Category.store_id == current_user.store_id))
    cat = result.scalar_one_or_none()
    if not cat:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Category not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(cat, k, v)
    return CategoryOut.model_validate(cat)


# ═══════════════════════════════════════════════════════════════════════════════
# SUPPLIERS
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/suppliers", tags=["Suppliers"])
async def list_suppliers(
    page: int = Query(1, ge=1),
    page_size: int = Query(20),
    search: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select, func, or_
    from app.models.models import Supplier
    query = select(Supplier).where(Supplier.store_id == current_user.store_id, Supplier.is_active == True)
    if search:
        query = query.where(or_(Supplier.name.ilike(f"%{search}%"), Supplier.phone.ilike(f"%{search}%")))
    total = (await db.execute(select(func.count()).select_from(query.subquery()))).scalar()
    result = await db.execute(query.offset((page - 1) * page_size).limit(page_size))
    return {"items": [SupplierOut.model_validate(s) for s in result.scalars().all()], "total": total}


@router.post("/suppliers", response_model=SupplierOut, tags=["Suppliers"])
async def create_supplier(
    data: SupplierCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.models.models import Supplier
    from app.services.activity_service import log_activity
    supplier = Supplier(store_id=current_user.store_id, **data.model_dump())
    db.add(supplier)
    await db.flush()
    await log_activity(db, current_user.store_id, current_user.id, "supplier_created", "supplier", supplier.id, {"name": supplier.name})
    return SupplierOut.model_validate(supplier)


@router.patch("/suppliers/{supplier_id}", response_model=SupplierOut, tags=["Suppliers"])
async def update_supplier(
    supplier_id: UUID,
    data: SupplierUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select
    from app.models.models import Supplier
    from fastapi import HTTPException, status
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id, Supplier.store_id == current_user.store_id))
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Supplier not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(supplier, k, v)
    return SupplierOut.model_validate(supplier)


@router.delete("/suppliers/{supplier_id}", tags=["Suppliers"])
async def delete_supplier(
    supplier_id: UUID,
    current_user: User = Depends(require_owner),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select
    from app.models.models import Supplier
    from fastapi import HTTPException, status
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id, Supplier.store_id == current_user.store_id))
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Supplier not found")
    supplier.is_active = False
    return {"message": "Supplier deleted"}


# ═══════════════════════════════════════════════════════════════════════════════
# PRODUCTS
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/products", tags=["Products"])
async def list_products(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    category_id: Optional[UUID] = Query(None),
    low_stock: bool = Query(False),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await ProductService.list_products(db, current_user.store_id, page, page_size, search, category_id, low_stock)

@router.get("/products/export", tags=["Products"])
async def export_products(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    df = await ProductService.export_products(
        db,
        current_user.store_id
    )

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:
        df.to_excel(
            writer,
            index=False,
            sheet_name="Inventory"
        )

    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            "attachment; filename=inventory_export.xlsx"
        }
    )   

@router.post("/products/import", tags=["Products"])
async def import_products(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from decimal import Decimal, InvalidOperation
    from fastapi import HTTPException, status
    from sqlalchemy import select

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Please upload an .xlsx inventory workbook.",
        )

    try:
        workbook = load_workbook(BytesIO(await file.read()), data_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded file is not a valid Excel workbook.",
        ) from exc

    sheet = workbook.active
    raw_headers = {
        str(value).strip().lower().replace("_", " ").replace("-", " "): index
        for index, value in enumerate(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        if value is not None
    }
    # Support both the application's exported workbook and a conventional
    # simple template (Name, SKU, Quantity, Purchase Price, Selling Price).
    header_aliases = {
        "product name": ("product name", "name", "product", "item name"),
        "sku": ("sku", "product sku", "sku / barcode", "sku/barcode"),
        "quantity": ("quantity", "qty", "stock", "stock quantity"),
        "purchase price": ("purchase price", "cost price", "cost", "unit cost", "price"),
        "selling price": ("selling price", "sale price", "retail price", "unit price", "price"),
        "barcode": ("barcode", "bar code"),
        "min stock": ("min stock", "minimum stock", "reorder level"),
    }
    headers = {
        canonical: next((raw_headers[alias] for alias in aliases if alias in raw_headers), None)
        for canonical, aliases in header_aliases.items()
    }

    # If no recognised headers are present, preserve compatibility with the
    # original five-column template: Name, SKU, Quantity, Purchase, Selling.
    if not any(index is not None for index in headers.values()):
        headers.update({
            "product name": 0,
            "sku": 1,
            "quantity": 2,
            "purchase price": 3,
            "selling price": 4,
        })
        first_data_row = 1
    else:
        first_data_row = 2

    # Product name, SKU and quantity are essential. Price columns are optional:
    # quantities can be imported first and priced later from the product editor.
    required_headers = {"product name", "sku", "quantity"}
    missing_headers = {header for header in required_headers if headers[header] is None}
    if missing_headers:
        return {
            "message": "No products were imported.",
            "imported": 0,
            "skipped": 1,
            "errors": [{
                "row": 1,
                "reason": "Missing required column(s): " + ", ".join(sorted(missing_headers)) + ".",
            }],
        }

    imported = 0
    restored = 0
    skipped: list[dict[str, str | int]] = []
    seen_skus: set[str] = set()
    missing_price_columns = [
        header for header in ("purchase price", "selling price")
        if headers[header] is None
    ]

    def value_at(row: tuple, header: str, default=None):
        index = headers.get(header)
        return row[index] if index is not None and index < len(row) else default

    for row_number, row in enumerate(sheet.iter_rows(min_row=first_data_row, values_only=True), start=first_data_row):
        if not any(value is not None and str(value).strip() for value in row):
            continue

        name = str(value_at(row, "product name") or "").strip()
        sku = str(value_at(row, "sku") or "").strip()
        try:
            quantity = int(value_at(row, "quantity"))
            purchase_price = Decimal(str(value_at(row, "purchase price", 0) or 0))
            selling_price = Decimal(str(value_at(row, "selling price", 0) or 0))
            min_stock = int(value_at(row, "min stock", 5) or 5)
        except (TypeError, ValueError, InvalidOperation):
            skipped.append({"row": row_number, "reason": "Quantity, prices, or minimum stock are invalid."})
            continue

        if not name or not sku:
            skipped.append({"row": row_number, "reason": "Product Name and SKU are required."})
            continue
        if quantity < 0 or purchase_price < 0 or selling_price < 0 or min_stock < 0:
            skipped.append({"row": row_number, "reason": "Quantity, prices, and minimum stock cannot be negative."})
            continue
        if sku in seen_skus:
            skipped.append({"row": row_number, "reason": f"Duplicate SKU '{sku}' in this workbook."})
            continue

        seen_skus.add(sku)
        existing = await db.scalar(select(Product).where(Product.sku == sku))
        if existing:
            # Deleted products are soft-deleted, so they remain in the SKU
            # index but are absent from the Inventory list and exports. An
            # import of that SKU should restore it instead of hiding it again.
            if existing.store_id == current_user.store_id and not existing.is_active:
                barcode = value_at(row, "barcode")
                existing.name = name
                existing.barcode = str(barcode).strip() if barcode is not None and str(barcode).strip() else None
                existing.quantity = quantity
                existing.purchase_price = purchase_price
                existing.selling_price = selling_price
                existing.min_stock = min_stock
                existing.is_active = True
                imported += 1
                restored += 1
                continue

            scope = "this store" if existing.store_id == current_user.store_id else "another store"
            skipped.append({"row": row_number, "reason": f"SKU '{sku}' already exists in {scope}."})
            continue

        barcode = value_at(row, "barcode")
        product = Product(
            store_id=current_user.store_id,
            name=name,
            sku=sku,
            barcode=str(barcode).strip() if barcode is not None and str(barcode).strip() else None,
            quantity=quantity,
            selling_price=selling_price,
            purchase_price=purchase_price,
            min_stock=min_stock,
        )
        db.add(product)
        await db.flush()

        if quantity:
            db.add(StockTransaction(
                store_id=current_user.store_id,
                product_id=product.id,
                user_id=current_user.id,
                type=TransactionType.STOCK_IN,
                quantity=quantity,
                cost_per_unit=purchase_price,
                remarks="Imported via Excel",
            ))
        imported += 1

    message = f"Imported {imported} product(s)."
    if restored:
        message += f" Restored {restored} previously deleted product(s)."
    if missing_price_columns:
        message += " Missing prices were set to ₹0.00; update them from the product editor."

    return {
        "message": message,
        "imported": imported,
        "skipped": len(skipped),
        "errors": skipped,
    }

@router.get("/products/barcode/{barcode}", tags=["Products"])
async def get_by_barcode(
    barcode: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    product = await ProductService.get_by_barcode(db, current_user.store_id, barcode)
    if not product:
        from fastapi import HTTPException, status
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")
    return ProductOut.model_validate(product)


@router.get("/products/{product_id}", response_model=ProductOut, tags=["Products"])
async def get_product(
    product_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await ProductService.get_product(db, current_user.store_id, product_id)


@router.post("/products", response_model=ProductOut, tags=["Products"])
async def create_product(
    data: ProductCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    product = await ProductService.create_product(db, current_user.store_id, current_user.id, data)
    await manager.broadcast_to_store(str(current_user.store_id), "product_created", {"id": str(product.id), "name": product.name})
    return ProductOut.model_validate(product)


@router.patch("/products/{product_id}", response_model=ProductOut, tags=["Products"])
async def update_product(
    product_id: UUID,
    data: ProductUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    product = await ProductService.update_product(db, current_user.store_id, current_user.id, product_id, data)
    return ProductOut.model_validate(product)


@router.delete("/products/{product_id}", tags=["Products"])
async def delete_product(
    product_id: UUID,
    current_user: User = Depends(require_owner),
    db: AsyncSession = Depends(get_db),
):
    await ProductService.delete_product(db, current_user.store_id, current_user.id, product_id)
    return {"message": "Product deleted"}


@router.get("/products/{product_id}/history", tags=["Products"])
async def product_history(
    product_id: UUID,
    page: int = Query(1),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await ProductService.get_product_history(db, current_user.store_id, product_id, page)


# ═══════════════════════════════════════════════════════════════════════════════
# STOCK
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/stock/in", tags=["Stock"])
async def stock_in(
    data: StockInCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    txn = await StockService.stock_in(db, current_user.store_id, current_user.id, data)
    # Fetch updated product for broadcast
    product = await ProductService.get_product(db, current_user.store_id, data.product_id)
    await manager.broadcast_to_store(
        str(current_user.store_id), "stock_update",
        {"product_id": str(data.product_id), "quantity": product.quantity, "type": "stock_in"}
    )
    return {"id": str(txn.id), "message": "Stock added successfully"}


@router.post("/stock/out", tags=["Stock"])
async def stock_out(
    data: StockOutCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    txn = await StockService.stock_out(db, current_user.store_id, current_user.id, data)
    product = await ProductService.get_product(db, current_user.store_id, data.product_id)
    await manager.broadcast_to_store(
        str(current_user.store_id), "stock_update",
        {"product_id": str(data.product_id), "quantity": product.quantity, "type": "stock_out"}
    )
    return {"id": str(txn.id), "message": "Stock removed successfully"}


@router.get("/stock/ledger", tags=["Stock"])
async def get_ledger(
    page: int = Query(1, ge=1),
    page_size: int = Query(20),
    transaction_type: Optional[str] = Query(None),
    product_id: Optional[UUID] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await StockService.get_ledger(db, current_user.store_id, page, page_size, transaction_type, product_id)


# ═══════════════════════════════════════════════════════════════════════════════
# AUDIT
# ═══════════════════════════════════════════════════════════════════════════════

@router.post("/audits", tags=["Audit"])
async def create_audit(
    data: AuditCreate,
    current_user: User = Depends(require_owner),
    db: AsyncSession = Depends(get_db),
):
    audit = await AuditService.create_audit(db, current_user.store_id, current_user.id, data)
    product = await ProductService.get_product(db, current_user.store_id, data.product_id)
    await manager.broadcast_to_store(
        str(current_user.store_id), "audit_complete",
        {"audit_id": str(audit.id), "product_name": product.name, "difference": audit.difference}
    )
    return AuditOut.model_validate(audit)


@router.get("/audits", tags=["Audit"])
async def list_audits(
    page: int = Query(1),
    page_size: int = Query(20),
    product_id: Optional[UUID] = Query(None),
    current_user: User = Depends(require_owner),
    db: AsyncSession = Depends(get_db),
):
    return await AuditService.list_audits(db, current_user.store_id, page, page_size, product_id)


@router.get("/audits/discrepancies", tags=["Audit"])
async def discrepancy_report(
    current_user: User = Depends(require_owner),
    db: AsyncSession = Depends(get_db),
):
    return await AuditService.get_discrepancy_report(db, current_user.store_id)


# ═══════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/dashboard/stats", tags=["Dashboard"])
async def dashboard_stats(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await DashboardService.get_stats(db, current_user.store_id)


@router.get("/dashboard/stock-movement", tags=["Dashboard"])
async def stock_movement_chart(
    days: int = Query(30, ge=7, le=365),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await DashboardService.get_stock_movement_chart(db, current_user.store_id, days)


@router.get("/dashboard/categories", tags=["Dashboard"])
async def category_distribution(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await DashboardService.get_category_distribution(db, current_user.store_id)


@router.get("/dashboard/movers", tags=["Dashboard"])
async def fast_slow_movers(
    days: int = Query(30),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await DashboardService.get_fast_slow_movers(db, current_user.store_id, days)


# ═══════════════════════════════════════════════════════════════════════════════
# NOTES
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/notes", tags=["Notes"])
async def list_notes(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select
    from app.models.models import Note
    result = await db.execute(select(Note).where(Note.store_id == current_user.store_id).order_by(Note.created_at.desc()))
    return result.scalars().all()


@router.post("/notes", tags=["Notes"])
async def create_note(
    data: NoteCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from app.models.models import Note
    note = Note(store_id=current_user.store_id, user_id=current_user.id, **data.model_dump())
    db.add(note)
    await db.flush()
    return note


@router.patch("/notes/{note_id}", tags=["Notes"])
async def update_note(
    note_id: UUID,
    data: NoteUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select
    from app.models.models import Note
    from fastapi import HTTPException, status
    result = await db.execute(select(Note).where(Note.id == note_id, Note.store_id == current_user.store_id))
    note = result.scalar_one_or_none()
    if not note:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Note not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(note, k, v)
    return note


@router.delete("/notes/{note_id}", tags=["Notes"])
async def delete_note(
    note_id: UUID,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select, delete
    from app.models.models import Note
    from fastapi import HTTPException, status
    result = await db.execute(select(Note).where(Note.id == note_id, Note.store_id == current_user.store_id))
    note = result.scalar_one_or_none()
    if not note:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Note not found")
    await db.delete(note)
    return {"message": "Note deleted"}


# ═══════════════════════════════════════════════════════════════════════════════
# ACTIVITY LOGS & NOTIFICATIONS
# ═══════════════════════════════════════════════════════════════════════════════

@router.get("/activity-logs", tags=["Logs"])
async def activity_logs(
    page: int = Query(1),
    page_size: int = Query(50),
    current_user: User = Depends(require_owner),
    db: AsyncSession = Depends(get_db),
):
    return await ActivityService.get_logs(db, current_user.store_id, page, page_size)


@router.get("/notifications", tags=["Notifications"])
async def get_notifications(
    unread_only: bool = Query(False),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    return await NotificationService.get_user_notifications(db, current_user.id, unread_only)


@router.post("/notifications/mark-read", tags=["Notifications"])
async def mark_notifications_read(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    await NotificationService.mark_all_read(db, current_user.id)
    return {"message": "All notifications marked as read"}


# ═══════════════════════════════════════════════════════════════════════════════
# WEBSOCKET
# ═══════════════════════════════════════════════════════════════════════════════

@router.websocket("/ws/{store_id}")
async def websocket_endpoint(websocket: WebSocket, store_id: str, token: str = ""):
    try:
        payload = decode_token(token)
        user_id = payload.get("sub", "unknown")
    except Exception:
        await websocket.close(code=4001, reason="Invalid token")
        return

    await manager.connect(websocket, store_id, user_id)
    try:
        while True:
            data = await websocket.receive_text()
            # Echo heartbeat
            if data == "ping":
                await websocket.send_text("pong")
    except WebSocketDisconnect:
        await manager.disconnect(websocket, store_id)
